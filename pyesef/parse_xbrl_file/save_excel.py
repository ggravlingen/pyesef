"""Save to Excel."""

from __future__ import annotations

from enum import StrEnum
import logging
import os
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from pandas import ExcelWriter

from pyesef.const import PATH_PROJECT_ROOT

if TYPE_CHECKING:
    from pyesef.parse_xbrl_file.read_and_save_filings import ReadFiling


class DataSheetName(StrEnum):
    """Represent data sheet names."""

    DATA = "Data"
    DEFINITIONS = "Definitions"


class SaveToExcel:
    """
    Class to save data to Excel.

    Appends the data if the file exists.
    """

    TEMPLATE_OUTPUT_PATH_EXCEL = os.path.join(PATH_PROJECT_ROOT, "output.xlsx")

    def __init__(self, parent: ReadFiling, df_to_save: pd.DataFrame) -> None:
        """Init class."""
        self.parent = parent
        self.df_to_save = df_to_save
        self.workbook: Workbook | None = None

        if df_to_save.empty:
            self.parent.cntlr.addToLog(
                "Empty output dataframe. Output file not saved.",
                level=logging.WARNING,
            )
            return

        self.main()

    def main(self) -> None:
        """Run program."""
        startrow = 0
        mode = "w"
        if_sheet_exists = None

        if os.path.exists(self.TEMPLATE_OUTPUT_PATH_EXCEL):
            reader = pd.read_excel(self.TEMPLATE_OUTPUT_PATH_EXCEL)
            startrow = reader.shape[0] + 1
            mode = "a"
            if_sheet_exists = "overlay"

        with pd.ExcelWriter(
            path=self.TEMPLATE_OUTPUT_PATH_EXCEL,
            engine="openpyxl",
            mode=mode,  # type: ignore[arg-type]
            if_sheet_exists=if_sheet_exists,  # type: ignore[arg-type]
            engine_kwargs={},
        ) as writer:
            self.save(writer=writer, startrow=startrow)

    def save(self, writer: ExcelWriter, startrow: int) -> None:
        """Save file."""
        to_excel_args = {
            "excel_writer": writer,
            "index": False,
            "sheet_name": DataSheetName.DATA.value,
            "freeze_panes": (1, 0),
            "startcol": 0,
            "startrow": startrow,
        }
        if startrow > 0:
            to_excel_args["header"] = None

        self.df_to_save.to_excel(**to_excel_args)  # type:ignore[arg-type]

        self._add_data_auto_filter(writer=writer)
        self._add_data_sheet_styling(writer=writer)
        self._save_definitions(writer=writer)

    def _add_data_auto_filter(self, writer: ExcelWriter) -> None:
        """Add auto filter to data sheet."""
        worksheet: Worksheet = writer.sheets[DataSheetName.DATA.value]

        worksheet.auto_filter.ref = (  # Enable autofilter
            "A1:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
        )

    def get_named_style_list(self, writer: ExcelWriter) -> list[str]:
        """Return a list of named styles."""
        return [
            style.name for style in writer.book._named_styles  # type: ignore # pylint: disable=protected-access
        ]

    def _add_data_sheet_styling(self, writer: ExcelWriter) -> None:
        """Add styling to data sheet."""
        worksheet: Worksheet = writer.sheets[DataSheetName.DATA.value]

        name_styles_name_list = self.get_named_style_list(writer=writer)

        # Add date style
        date_style_obj = NamedStyle(name="DateCol", number_format="yyyy-mm-dd")
        date_style: NamedStyle | str
        if "DateCol" in name_styles_name_list:
            date_style = "DateCol"
        else:
            date_style = date_style_obj

        # Apply the date style to the entire column
        for cell in worksheet["A"]:
            cell.style = date_style

        # Add int style
        int_style_obj = NamedStyle(name="IntStyle", number_format="0")
        int_style: NamedStyle | str
        if "IntStyle" in name_styles_name_list:
            int_style = "IntStyle"
        else:
            int_style = int_style_obj

        # Apply the integer style to the value column (column 'B')
        for cell in worksheet["G"]:
            cell.style = int_style

    def _save_definitions(self, writer: ExcelWriter) -> None:
        """Save definitions sheet."""
        if (
            not self.parent.definitions.empty
            and DataSheetName.DEFINITIONS.value not in writer.sheets
        ):
            self.parent.definitions.to_excel(
                writer,
                index=False,
                sheet_name=DataSheetName.DEFINITIONS.value,
                freeze_panes=(1, 0),
            )
